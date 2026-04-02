import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from datetime import datetime

st.set_page_config(page_title="CRM Data Import Tool", layout="centered")
st.title("🛠️ Import Data Processing Tool")

st.markdown("---")
source_file = st.file_uploader("1. Please upload Source data file (Excel/CSV)", type=["xlsx", "csv"])
template_file = st.file_uploader("2. Please upload Template file (Excel)", type=["xlsx"])
global_country_code = st.text_input("3. Please input Country Code (Exp: VN, NP, AU,...):").strip().upper()

if st.button("🚀 Start processing"):
    if not source_file or not template_file or not global_country_code:
        st.warning("⚠️ Please upload both files and enter the country code before running.!")
    else:
        try:
            with st.spinner('Processing data...'):
                # Đọc data
                if source_file.name.endswith('.csv'):
                    source_df = pd.read_csv(source_file)
                else:
                    source_df = pd.read_excel(source_file)
                
                # Đọc template
                wb = load_workbook(template_file)
                ws = None
                header_row = 1
                template_headers = {}
                core_keywords = ['品牌', '库存状态', 'VIN码', 'VIN号', '国家/地区', '仓库']
                
                for sheet_name in wb.sheetnames:
                    temp_ws = wb[sheet_name]
                    found_header = False
                    for r in range(1, 9):
                        temp_headers = {}
                        for cell in temp_ws[r]:
                            if cell.value:
                                c_name = str(cell.value).replace('*', '').replace(' ', '').replace('\n', '').replace('\r', '').strip()
                                temp_headers[c_name] = cell.column
                        
                        matches = sum(1 for k in core_keywords if k in temp_headers)
                        if matches >= 2:
                            ws = temp_ws
                            header_row = r
                            template_headers = temp_headers
                            found_header = True
                            break
                    if found_header:
                        break

                if not ws:
                    st.error("❌ No sheets containing valid vehicle data columns were found in the Template file.")
                    st.stop()

                # Bắt đầu điền từ dòng 9 (bỏ qua dòng 8)
                start_row = 9 
                missing_sap_count = 0

                for index, row in source_df.iterrows():
                    current_row = start_row + index

                    # Default values
                    defaults = {'所属组织': 'ASIA0000', '组织': 'ASIA0000', '品牌': '比亚迪', '库存状态': 2, '仓库': '0922', '车辆状态': 1}
                    for col_name, val in defaults.items():
                        if col_name in template_headers:
                            ws.cell(row=current_row, column=template_headers[col_name], value=val)

                    # Direct mappings
                    direct_map = {'车辆备注': '备注', 'VIN码': 'VIN号', 'vin年份': '生产年份'}
                    if 'VIN码' not in template_headers and 'VIN号' in template_headers:
                        direct_map['VIN号'] = 'VIN号'
                        del direct_map['VIN码']

                    for tmpl_col, src_col in direct_map.items():
                        if tmpl_col in template_headers and src_col in source_df.columns:
                            val = row[src_col]
                            if pd.notna(val):
                                ws.cell(row=current_row, column=template_headers[tmpl_col], value=val)

                    # Country
                    if '国家/地区' in template_headers:
                        ws.cell(row=current_row, column=template_headers['国家/地区'], value=global_country_code)

                    # SAP Code logic
                    sap_code = ""
                    if '销售料号' in source_df.columns and pd.notna(row['销售料号']):
                        sap_code = str(row['销售料号']).strip()
                    elif 'SAP号' in source_df.columns and pd.notna(row['SAP号']):
                        # Trên web tạm thời mượn luôn SAP号 nếu thiếu 销售料号
                        sap_code = str(row['SAP号']).strip() 
                    else:
                        missing_sap_count += 1

                    if 'SAP销售料号' in template_headers:
                        ws.cell(row=current_row, column=template_headers['SAP销售料号'], value=sap_code)
                    elif '销售料号' in template_headers:
                        ws.cell(row=current_row, column=template_headers['销售料号'], value=sap_code)

                # Lưu vào bộ nhớ đệm (BytesIO) để tải về qua web
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success(f"✅ Successfully processed {len(source_df)} rows!")
                if missing_sap_count > 0:
                    st.warning(f"⚠️ There are {missing_sap_count} lines indicating the SAP code was not found. Please double-check your output file.")

                date_str = datetime.now().strftime("%y%m%d")
                output_filename = f"Data_Import_{global_country_code}_{date_str}.xlsx"

                st.download_button(
                    label="📥 Download the results file to your computer",
                    data=output,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"An error occurred during processing: {e}")
